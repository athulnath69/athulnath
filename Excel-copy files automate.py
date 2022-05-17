# -*- coding: utf-8 -*-
"""
Created on Mon May 16 11:54:44 2022

@author: athul
"""

# importing openpyxl module
import openpyxl as xl
import os
#import time
import subprocess
# opening the source excel file
subprocess.Popen([r"C:\Users\athul\Downloads\pythontest.xlsx"],shell=True)
filename =r"C:\Users\athul\Downloads\pythontest.xlsx"
wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[0]

# opening the destination excel file
filename1 =r"C:\Users\athul\Downloads\copydata.xlsx"
wb2 = xl.load_workbook(filename1)
ws2 = wb2.active

# calculate total number of rows and
# columns in source excel file
#creating vairables
mr = ws1.max_row
mc = ws1.max_column

# copying the cell values from source
# excel file to destination excel file
for i in range (1, mr + 1):
	for j in range (1, mc + 1):
		# reading cell value from source excel file
		c = ws1.cell(row = i, column = j)

		# writing the read value to destination excel file
		ws2.cell(row = i, column = j).value = c.value


# saving the destination excel file
wb2.save(str(filename1))
#open any workbook 
#subprocess.Popen([r"C:\Users\athul\Downloads\pythontest.xlsx"],shell=True)
os.system(r"C:\Users\athul\Downloads\copydata.xlsx")
#open anyprograme

#r"C:\Users\athul\Downloads\copydata.xlsx"
#time.sleep(5)